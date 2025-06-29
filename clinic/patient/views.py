from django.shortcuts import render, redirect, get_object_or_404 
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.utils.translation import gettext as _
import magic  # python-magic-bin for mimetype detection 

from django.contrib.auth.models import User
from django.http import HttpRequest, HttpResponse

from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

import openpyxl
from openpyxl.styles import PatternFill, Font

from django.template.loader import render_to_string
from xhtml2pdf import pisa


from account.views import login_required_with_message
from django.contrib import messages
from datetime import datetime
from django.utils import timezone

from account.models import Profile, MedicalInfo, ActivityLog, Conversation, Message, Calls
from doctor.models import DoctorProfile, AppointmentDateSlot, AppointmentTimeSlot
from patient.models import *

from account.utils import log_action
from django.urls import reverse

from django.core.serializers.json import DjangoJSONEncoder
import json
from django.views.decorators.csrf import csrf_exempt

from home.send_email import send_custom_email
import uuid
from django.conf import settings
DOMAIN_NAME = settings.DOMAIN_NAME

# Constants
ALLOWED_FILE_TYPES_APPOINTMENT = [
            # Documents
            'application/pdf',
            'application/msword', # .doc
            'application/vnd.openxmlformats-officedocument.wordprocessingml.document', # .docx
            'text/plain', # .txt

             # Images
            'image/jpeg',
            'image/png',
            'image/jpg',
            'image/gif',
            'image/webp',

             # Audio
            'audio/mpeg',  # .mp3
            'audio/wav',
            'audio/ogg',

            # Video
            'video/mp4',
            'video/ogg',
            'video/webm'
        ]

ALLOWED_FILE_TYPES_DOC = [
    'application/pdf',
    'image/jpeg',
    'image/png',
    'image/jpg',
    'application/msword',
    'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
]
MAX_FILE_SIZE_MB = 5

def is_valid_file(uploaded_file, allowed_file_types=ALLOWED_FILE_TYPES_DOC, max_file_size_mb=MAX_FILE_SIZE_MB):
    if not uploaded_file:
        return False, _("No file uploaded.")

    # Check size
    if uploaded_file.size > max_file_size_mb * 1024 * 1024:
        return False, _(f"File size exceeds {max_file_size_mb} MB.")

    # Check file type using content sniffing
    file_type = magic.from_buffer(uploaded_file.read(2048), mime=True)
    uploaded_file.seek(0)  # reset file pointer after reading
    if file_type not in allowed_file_types:
        return False, _("Unsupported file type.")

    return True, None


def redirect_to_role_dashboard(request: HttpRequest):
    """Redirect to the specific role dashboard."""
    profile: Profile = request.user.profile
    if profile.role == 'doctor':
        return redirect('doctor:doctor_dashboard')
    return redirect('patient:patientDashboard')





# --------------------------------------- Rendering Pages ------------------------------------------------------------

@login_required_with_message(login_url='account:login', message="You need to log in to Access this Page.", only=['patient'])
def patientDashboard(request: HttpRequest):
    """Patient dashboard view."""

    profile: Profile = request.user.profile

    # Get all schedules from all prescriptions for this profile
    all_schedules = PrescriptionSchedule.objects.filter(
        prescription__profile=profile
    ).order_by('-time')

    context = {
        'profile': profile,
        'todayMedications': all_schedules
    }

    return render(request, 'pages/patient/dashboard.html', context)

@login_required_with_message(login_url='account:login', message="You need to log in to View Your Appointments.", only=['patient'])
def viewAppointment(request: HttpRequest):
    profile = Profile.objects.get(user=request.user)
    qs = Appointment.objects.filter(profile=profile).order_by('-created_at')

    # pull per_page from GET, default 10
    per_page = int(request.GET.get('per_page', 10))
    paginator = Paginator(qs, per_page)
    page_num = request.GET.get('page')

    try:
        page_obj = paginator.page(page_num)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    # define your entries‑per‑page options here
    per_page_options = [10, 25, 50, 100]

    context = {
        'profile': profile,
        'appointments': page_obj,
        'paginator': paginator,
        'per_page': per_page,
        'per_page_options': per_page_options,
    }
    return render(request, 'pages/patient/view_appointment.html', context)


@login_required_with_message(login_url='account:login', message="You need to log in to Get Excel Files.")
def export_appointments_excel(request):
    profile: Profile = request.user.profile

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Appointments"

    # Add a description/info row at the top
    ws.append([
        f"Exported Appointments for: {profile.user.get_full_name()} (User ID: {profile.user.username})"
    ])
    ws.append([
        f"Export Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ])
    ws.append([])  # Empty row for spacing
    ws.append([])  # Empty row for spacing

    # Header
    # Define a fill color (light blue) and bold font for the header
    header_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    header_font = Font(bold=True)

    ws.append([
        "ID", "Doctor", "Specialization", "Appointment Type", "Date", "Time",
        "Reason For Appointment", "Status", "Cancelled By", "Cancel Reason"
    ])

    # Apply the fill and font to the header row
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = header_font

    appointments  = Appointment.objects.filter(profile=profile).order_by('-created_at')
    for appt in appointments:
        cancelled_by = appt.cancled_by.user.first_name if appt.status == 'cancelled' and appt.cancled_by else "Null"
        cancel_reason = appt.cancel_reason if appt.status == 'cancelled' else "Null"
        ws.append([
            str(appt.uuid),
            appt.doctor.profile.user.first_name,
            appt.doctor.specialization,
            appt.appointment_type,
            appt.appointment_date.strftime('%d %b %Y'),
            f"{appt.appointment_time_str}",
            appt.reason,
            appt.status,
            cancelled_by,
            cancel_reason,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=appointments.xlsx"
    wb.save(response)
    return response



@login_required_with_message(login_url='account:login', message="You need to log in to Delete, Edit Appointments.")
def appoinemtCancle_Edit(request: HttpRequest, apot_id: uuid, status: str):
    try:
        appointment = get_object_or_404(Appointment, uuid=apot_id)
        profile: Profile = Profile.objects.get(user=request.user)

        common_reason = request.POST.get('common_reason', '').strip()
        custom_reason = request.POST.get('custom_reason', '').strip()

        print(f"Common reason: {common_reason}, Custom reason: {custom_reason}")

        if common_reason or custom_reason:
            appointment.cancel_reason = common_reason if common_reason != 'other' else f"Custom Reason: {custom_reason}"
            appointment.cancled_by = request.user.profile
            appointment.save()

        if request.method == 'POST':
            if status == 'cancel':
                if appointment.status != "completed":
                    if appointment.time_slot:
                        appointment.time_slot.status = 'available'
                        appointment.time_slot.save()
                    appointment.status = 'cancelled'
                    appointment.save()

                    # send email notification to doctor
                    send_custom_email(
                        subject="Appointment Cancelled",
                        message=(
                            f"Dear {appointment.doctor.profile.user.get_full_name()},\n\n"
                            f"Your appointment with {profile.user.get_full_name()} on {appointment.appointment_date} at {appointment.appointment_time_str} has been cancelled.\n"
                            f"Reason: {appointment.cancel_reason}\n\n"
                            f"Thank you for your understanding.\n\n"
                            f"Best regards,\nNCMS Team"
                        ),
                        recipient_list=[appointment.doctor.profile.user.email]
                    )

                    # Log the action
                    log_action(
                        profile=profile,
                        action='CANCEL_APPT',
                        title=_("Appointment Cancelled"),
                        description=_("Appointment with Dr. {} on {} at {} has been cancelled.").format(
                            appointment.doctor.profile.user.get_full_name(),
                            appointment.appointment_date,
                            appointment.appointment_time_str
                        ),
                        obj=appointment
                    )
                    messages.success(request, _("Appointment cancelled successfully."))
                else:
                    messages.error(request, _("Cannot cancel a completed appointment."))
            elif status == 'edit':
                # Logic for editing the appointment can be added here
                pass
            return redirect('patient:viewAppointment')
    except Exception as e:
        print(f"Error: {e}")
        messages.error(request, _("An error occurred while processing your request."))
        return redirect('patient:viewAppointment')

@login_required_with_message(login_url='account:login', message="You need to log in to Book an appointment.")
def BookAppointment(request: HttpRequest):
    """Doctor booking page."""

    if request.method == 'GET':
        # Fetch the user's profile information
        profile: Profile = request.user.profile
        doctors: DoctorProfile = DoctorProfile.objects.all()

        doctor_data = []
        for doc in doctors:
            date_slots = AppointmentDateSlot.objects.filter(doctor=doc)
            date_json = {}
            for each_date_slot in date_slots:
                date_str = each_date_slot.date.strftime('%Y-%m-%d')
                times_slots = AppointmentTimeSlot.objects.filter(appointment_date_slot=each_date_slot)

                for each_time_slot in times_slots:

                    time_str = f"{each_time_slot.from_time.strftime('%H:%M')} -- {each_time_slot.to_time.strftime('%H:%M')}"
                    all_selected_types = list(each_time_slot.appointment_type)

                    if each_time_slot.status not in ['booked', 'unavailable', 'break']:
                        if date_str not in date_json:
                            date_json[date_str] = {}
                        date_json[date_str][time_str] = [each_time_slot.id, each_time_slot.duration, all_selected_types]

            doctor_data.append({
                'id': doc.id,
                'name': f"Dr. {doc.profile.user.first_name}",
                'specialty': doc.get_specialization_display(),
                'experience': doc.experience_years,
                'rating': float(doc.star_rating) if doc.star_rating else None,
                'reviews': doc.total_reviews,
                'fees': float(doc.fees),
                'image': doc.profile.profile_pic.url,
                'availability': date_json,
            })


        context = {
            'profile': profile,
            'doctors': doctors,
            'doctor_data_json': json.dumps(doctor_data, cls=DjangoJSONEncoder)
        }

        return render(request, 'pages/patient/book_appointment.html', context)

    elif request.method == 'POST':
        try:
            # Fetch the user's profile information
            profile: Profile = Profile.objects.get(user=request.user)

            # Get the selected doctor and appointment details from the form
            doctor_id = request.POST.get('doctor_id')
            appointment_type = request.POST.get('appointment_type')
            appointment_date = request.POST.get('appointment_date')
            appointment_time = request.POST.get('appointment_time')
            appointment_time_slot_id = request.POST.get('appointment_time_id')
            appointment_reason = request.POST.get('appointment_reason', '').strip()

            appointment_file = request.FILES.get('appointment_file')

            # Fetch the doctor and date slot objects
            doctor: DoctorProfile = get_object_or_404(DoctorProfile, id=doctor_id)
            time_slot_instance: AppointmentTimeSlot = get_object_or_404(AppointmentTimeSlot, id=appointment_time_slot_id)

            # The user’s Profile will be created automatically via signals

            time_slot_instance.status = 'booked'
            time_slot_instance.save()   
            appointment_date_foramt = datetime.strptime(appointment_date, '%Y-%m-%d').date()
            # Create a new appointment
            appointment = Appointment.objects.create(
                profile=profile,
                doctor=doctor,
                time_slot=time_slot_instance,

                appointment_type=appointment_type,
                appointment_date=appointment_date_foramt,
                appointment_time_str=appointment_time,

                reason = appointment_reason,
                status='pending'
            )

            # build the connections conversation
            # Check if a conversation already exists between the patient and doctor
            existing_conversation: Conversation = Conversation.objects.filter(
                participants=profile
            ).filter(
                participants=doctor.profile
            ).distinct().first()
            
            # Only create new conversation if one doesn't exist
            if not existing_conversation:
                conversation = Conversation.objects.create(
                    uuid=uuid.uuid4(),
                    status='initiated',
                )
                conversation.participants.add(profile, doctor.profile)
                conversation.save()
                existing_conversation = conversation

            if appointment.appointment_type == 'online_consultation':
                # Create a new call record for the appointment
                call = Calls.objects.create(
                    uuid=uuid.uuid4(),
                    appointment=appointment,
                    connection=existing_conversation,
                    caller=profile,
                    receiver=doctor.profile,
                    last_req=timezone.now(),
                    status='requested'
                )
                call.save()

            if appointment_file:
                is_valid, error_msg = is_valid_file(appointment_file, ALLOWED_FILE_TYPES_APPOINTMENT, 20)
                if not is_valid:
                    messages.error(request, error_msg)
                    return JsonResponse({'error': error_msg})
                appointment.file = appointment_file
                appointment.save()
            
            # Send Email with appointment details
            send_custom_email(
                subject="Appointment Booked!",
                message=(
                    f"Dear {profile.user.get_full_name()},\n\n"
                    f"Your appointment has been booked successfully.\n\n"
                    f"Appointment Details:\n"
                    f"Booked By: {profile.user.get_full_name()} (Patient)\n"
                    f"Doctor: Dr. {doctor.profile.user.get_full_name()}\n"
                    f"Specialization: {doctor.get_specialization_display()}\n"
                    f"Date: {appointment_date}\n"
                    f"Time: {appointment_time}\n"
                    f"Type: {appointment_type.replace('_', ' ').title()}\n"
                    f"Reason: {appointment_reason or 'N/A'}\n\n"
                    f"Thank you for using NCMS.\n\n"
                    f"Best regards,\nNCMS Team"
                ),
                recipient_list=[profile.user.email, doctor.profile.user.email]
            )

            # Log the action
            log_action(
                profile=profile,
                action='BOOK_APPT',
                title=_("Appointment Scheduled"),
                description=_("Appointment booked with Dr. {} on {} at {}").format(
                    doctor.profile.user.get_full_name(),
                    appointment_date,
                    appointment_time
                ),
                obj=appointment
            )

                        #Send Email
            
            messages.success(request, _("Appointment booked successfully."))
            return JsonResponse({'success': True, 'redirect_url': reverse('patient:viewAppointment')})
        except Exception as e:
            messages.error(request, _("An error occurred while booking the appointment."))
            error_msg = str(e)
            return JsonResponse({'error': error_msg})

    return redirect('patient:bookAppointment')


# --------------------------------------- Document Management ------------------------------------------------------------
@login_required_with_message(login_url='account:login', message="You need to log in to View Your Files/ Document.", only=['patient'])
def ViewDocument(request: HttpRequest):
    path = reverse('patient:viewDocument')
    if request.method == 'POST':
        nick_name = request.POST.get('nick_name', '').strip()
        doc_type = request.POST.get('doc_type', '').strip()
        notes = request.POST.get('notes', '').strip()
        uploaded_file = request.FILES.get('document_file')

        if not nick_name or not doc_type:
            messages.error(request, "Please fill in all required fields.")
            return redirect(path+ '#add_new')

        is_valid, error_msg = is_valid_file(uploaded_file)
        if not is_valid:
            messages.error(request, error_msg)
            return redirect(path+'#add_new')
        try:
            profile: Profile = request.user.profile
            doc = Documents.objects.create(
                profile=profile,
                nick_name=nick_name,
                doc_type=doc_type,
                notes=notes,
                file=uploaded_file,
            )
            # Log the action
            log_action(
                profile=profile,
                action='UPLOAD_DOC',
                title=_("Document Uploaded"),
                description=_("Document '{}' uploaded.").format(nick_name),
                obj=doc
            )

            messages.success(request, "Document uploaded successfully.")
        except Exception as e:
            messages.error(request, "An error occurred while saving the document.")
            return redirect(path+'#add_new')


        return redirect('patient:viewDocument')
    


    profile: Profile = Profile.objects.get(user=request.user)
    documents: Documents = Documents.objects.filter(profile= profile).order_by('-created_at')
    context= {
        'profile': profile,
        'documents' : documents
    }
    return render(request, 'pages/patient/view_document.html', context)

@login_required_with_message(login_url='account:login', message="You need to log in to Update Your Files/ Document.")
def delete_document(request, doc_id):
    try:
        document = get_object_or_404(Documents, id=doc_id)
        
        # Optional: Only allow delete on POST
        if request.method == "POST":
            if document.file:
                document.file.delete(save=False)  # Delete file from storage
            document.delete()  # Delete record from DB
            messages.success(request, "Document deleted successfully.")

            log_action(
                profile=request.user.profile,
                action='DELETE_DOC',
                title=_("Document Deleted"),
                description=_("Document '{}' deleted.").format(document.nick_name),
                obj=document
            )


            return redirect('patient:viewDocument')  # Redirect to your document list page
    except Documents.DoesNotExist:
        messages.error(request, "Document not found.")
    except Exception as e:  
        messages.error(request, f"An error occurred: {e}")

    # If accessed via GET, redirect to same page (or show a confirm page optionally)
    return redirect('patient:viewDocument')



# video calls 
@login_required_with_message(login_url='account:login', message="You need to log in to Manage Video Calls", only=['patient'])
def view_v_call(request: HttpRequest):
    """Request a video call with a doctor."""
    if request.method != 'POST':
        profile: Profile = request.user.profile

        convs =  Conversation.objects.filter(
            participants=request.user.profile
        ).order_by('-created_at')
        for conversation in convs:
            # Get the other participant (not the current user)
            conversation.other_participant = conversation.participants.exclude(
                id=profile.id
            ).first()

    return render(request, 'pages/patient/list-v-call.html',{ 'conv': convs,}  )

@login_required_with_message(login_url='account:login', message="You need to log in to Send Calls Request", only=['patient'])
def send_req_calls(request: HttpRequest, convo_uuid: uuid):
    """Send a request for a video call."""
    try:
        profile: Profile = request.user.profile
        conversation: Conversation = get_object_or_404(Conversation, uuid=convo_uuid)

        # Ensure the user is part of the conversation
        if profile not in conversation.participants.all():
            messages.error(request, _("You are not part of this conversation."))
            return redirect('patient:view_v_call')

        # create a call object
        call = Calls.objects.create(
            uuid=uuid.uuid4(),
            connection=conversation,
            caller=profile,
            last_req=timezone.now(),
            status='requested',
            receiver=conversation.participants.exclude(id=profile.id).first()
        )

        Message.objects.create(
            conversation=conversation,
            sender=profile,
            content=f"Call Request from {profile.user.first_name}",
        )

        if call.receiver.role == "patient":
            #send Mail for patient
            send_custom_email(
                    subject=f"Call Request, From: DR. {call.caller.user.first_name}",
                    message=f"Hi, The Call was Requested. \n\nFrom: Dr. {call.caller.user.first_name}  \nTo: {call.receiver.user.first_name} \n\nPlz Get Free And Join a Call      \n\n\n#{DOMAIN_NAME}/p/join-v-call/{call.uuid}/ ",
                    recipient_list=[call.caller.user.email]
            )
        elif call.receiver.role == "doctor":
            #send Mail for Doctor
            send_custom_email(
                    subject=f"Call Request, From: {call.caller.user.first_name} ",
                    message=f"Hi, The Call was Requested. \n\nFrom: {call.caller.user.first_name}  \nTo: Dr.{call.receiver.user.first_name} \n\nPlz Get Free And Join a Call      \n\n\n#{DOMAIN_NAME}/d/join-v-call/{call.uuid}/ ",
                    recipient_list=[call.caller.user.email]
            )

        messages.success(request, _("Video call request sent successfully."))
        return redirect('patient:join_v_call', calls_uuid=call.uuid)
    except Exception as e:
        print(f"Error: {e}")
        messages.error(request, _("An error occurred while sending the video call request."))
        return redirect('patient:view_v_call')

@login_required_with_message(login_url='account:login', message="You need to log in to Join Video Calls", only=['patient'])
def join_v_call(request: HttpRequest, calls_uuid: uuid):
    profile: Profile = request.user.profile
    calls: Calls = get_object_or_404(Calls, uuid=calls_uuid)
    # Ensure the call exists, completed or cancelled calls cannot be joined
    if calls.status not in ['requested', 'active', 'ongoing']:
        messages.error(request, _("The call cannot be joined as it is either completed or cancelled."))
        return redirect('patient:view_v_call')


    conversation: Conversation = calls.connection

    # Ensure the user is part of the conversation
    if profile not in conversation.participants.all():
        messages.error(request, _("You are not part of this conversation."))
        return redirect('patient:view_v_call')
    
    conversation.other_participant = conversation.participants.exclude(
        id=profile.id
    ).first()

    is_caller = calls.caller == profile

    send_custom_email(
        subject=f"Call Request, From: {calls.caller.user.first_name}",
        message=f"Hi, The Call was Requested. \n\nFrom: {calls.caller.user.first_name}  \nTo: {calls.receiver.user.first_name} \n\nPlz Get Free And Join a Call      \n\n\n#{DOMAIN_NAME}/p/join-v-call/{calls.uuid}/ ",
        recipient_list=[calls.receiver.user.email]
    )

    return render(request, 'pages/patient/join-v-call.html', {'conversation': conversation,
                                                                'call_obj': calls,
                                                                'is_caller': is_caller,})

def waiting_room(request: HttpRequest, calls_uuid: uuid):
    """Join a video call with a specific conversation ID."""
    profile: Profile = request.user.profile
    calls: Calls = get_object_or_404(Calls, uuid=calls_uuid)
    conversation: Conversation = calls.connection

    # Ensure the user is part of the conversation
    if profile not in conversation.participants.all():
        messages.error(request, _("You are not part of this conversation."))
        return redirect('patient:view_v_call')
    

    conversation.other_participant = conversation.participants.exclude(
        id=profile.id
    ).first()
    
    return render(request, 'pages/patient/waiting-room.html', {'conversation': conversation,
                                                                'call_obj': calls,})


@login_required_with_message(login_url='account:login', message="You need to log in to view your Messages.", only=['patient'])
def message(request: HttpRequest):
    profile : Profile = request.user.profile
    
    conversations = Conversation.objects.filter(
        participants=profile
    )

    connected_paritcipants = []
    
    # Add additional data to each conversation
    for conversation in conversations:
        # Get the other participant (not the current user)
        conversation.other_participant = conversation.participants.exclude(
            id=profile.id
        ).first()

        if conversation.other_participant:
            connected_paritcipants.append(conversation.other_participant)
        
        # Get the last message
        conversation.last_message = conversation.messages.last()
        # Check if there are unread messages
        conversation.has_unread = conversation.messages.filter(
            read=False
        ).exclude(sender=profile).exists()

    # # Mark messages as read
    # conversation.messages.filter(
    #     read=False
    # ).exclude(sender=profile).update(read=True)

    # messages_list = conversation.messages.all().order_by('timestamp')
    # concept_data = {'active_conversation_id': conversation.id,
    #     'active_conversation': conversation,
    #     'message_lists': messages_list,}




    # Get all other not connected participants excluding the current user
    connected_ids = [p.id for p in connected_paritcipants]
    connected_ids.append(profile.id)
    if profile.role == 'doctor':
        not_connected_usr = Profile.objects.exclude(id__in=connected_ids).filter(role='patient')
    else:
        not_connected_usr = Profile.objects.exclude(id__in=connected_ids).filter(role='doctor')

    context = {
        'profile': profile,
        'conversations': conversations,
        'not_connected_usr': not_connected_usr,
        # 'active_conversation_id': conversation.id,
        # 'active_conversation': conversation,
        # 'message_lists': messages_list,
    }

    return render(request, 'pages/patient/message.html', context)


def req_conv(request: HttpRequest):
    """Create a new conversation with another user."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method.'}, status=405)

    profile: Profile = request.user.profile
    data = json.loads(request.body)
    other_username = data.get('userID')

    if not other_username:
        return JsonResponse({'error': 'Other user ID is required.'}, status=400)

    other_profile = get_object_or_404(Profile, user__username=other_username)

    if other_profile == profile:
        return JsonResponse({'error': 'You cannot start a conversation with yourself.'}, status=400)

    print(f"Profile: {profile}, Other Profile: {other_profile}")

    # Check if a conversation already exists between the two users
    # Find if a conversation exists with exactly these two participants (no more, no less)
    conversation = (
        Conversation.objects
        .filter(participants=profile)
        .filter(participants=other_profile)
        .distinct()
    )
    # Check if any conversation has exactly these two participants (and no more)
    for conv in conversation:
        if conv.participants.count() == 2:
            return JsonResponse({'error': 'Conversation already exists.'}, status=400)

    else:
        conversation = Conversation.objects.create(status='requested')
        conversation.participants.add(profile, other_profile)

    # Add both profiles to the conversation
    # conversation.participants.add(profile, other_profile)
    
    payload = {
        'conversation_id': conversation.id,
        'other_participant_name': other_profile.user.get_full_name(),
        'other_participant_pic': other_profile.profile_pic.url,
    }

    return JsonResponse({'success': True, 'payload': payload}, safe=False, json_dumps_params={'indent': 2})


def get_msg_list(request: HttpRequest, conversation_id: int):

    if request.method != 'GET':
        return JsonResponse({'error': 'Invalid request method.'}, status=405)
    """Fetch messages for a specific conversation."""

    profile: Profile = request.user.profile
    conversation = get_object_or_404(Conversation, id=conversation_id)

    # Ensure the user is part of the conversation
    if profile not in conversation.participants.all():
        return JsonResponse({'error': 'You are not part of this conversation.'}, status=403)

    # Fetch messages for the conversation
    messages_list = conversation.messages.all().order_by('timestamp')

    # Mark messages as read
    messages_list.filter(read=False).exclude(sender=profile).update(read=True)

    other_participant = conversation.participants.exclude(id=profile.id).first()

    all_msg = []
    for message in messages_list:
        local_dt = timezone.localtime(message.timestamp)
        timestamp = local_dt.strftime('%I:%M %p, %d %b %Y')
        all_msg.append({
            'msg_by_me': message.sender == profile,
            'id': message.id,
            'sender': message.sender.user.get_full_name(),
            'content': message.content,
            'timestamp': timestamp,
            'read': message.read,
        })

    payload = {
            'other_participant_name': other_participant.user.get_full_name(),
            'other_participant_pic' : other_participant.profile_pic.url,
            'conversation_id': conversation.id,
            'messages': all_msg,
    }

    return JsonResponse(payload, safe=False, json_dumps_params={'indent': 2})


def post_msg(request: HttpRequest):
    """Post a new message to the conversation."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method.'}, status=405)

    profile: Profile = request.user.profile
    data = json.loads(request.body)
    conversation_id = data.get('conversation_id')
    content = data.get('content', '').strip()

    if not content:
        return JsonResponse({'error': 'Message content cannot be empty.'}, status=400)

    conversation = get_object_or_404(Conversation, id=conversation_id)

    # Ensure the user is part of the conversation
    if profile not in conversation.participants.all():
        return JsonResponse({'error': 'You are not part of this conversation.'}, status=403)

    # Create the message
    message = Message.objects.create(
        sender=profile,
        conversation=conversation,
        content=content,
        read=False
    )

    messages_list = conversation.messages.all().order_by('timestamp')
    other_participant = conversation.participants.exclude(id=profile.id).first()
    all_msg = []
    for message in messages_list:
        local_dt = timezone.localtime(message.timestamp)
        timestamp = local_dt.strftime('%I:%M %p, %d %b %Y')
        all_msg.append({
            'msg_by_me': message.sender == profile,
            'id': message.id,
            'sender': message.sender.user.get_full_name(),
            'content': message.content,
            'timestamp': timestamp,
            'read': message.read,
        })

    payload = {
            'other_participant_name': other_participant.user.get_full_name(),
            'other_participant_pic' : other_participant.profile_pic.url,
            'conversation_id': conversation.id,
            'messages': all_msg,
    }

    return JsonResponse({'success': True, 'payload': payload}, safe=False, json_dumps_params={'indent': 2})



@login_required_with_message(login_url='account:login', message="You need to log in to access your Lab Reports.", only=['patient'])
def labReport(request: HttpRequest):
    profile: Profile = request.user.profile
    reports: LabReport = LabReport.objects.filter(patient_profile=profile)

    
    abnormal_reports = 0
    pending_reports = 0

    payload = []
    for each_report in reports:
        if each_report.status == 'abnormal':
                abnormal_reports+=1
        elif each_report.status == 'pending':
                pending_reports+=1

        params = []
        for each_params in each_report.parameters.all():
            params.append({
                'name':     each_params.parameter_name,
                'result':     each_params.result,
                'reference': each_params.reference_range,
                'status':   each_params.status.lower(),
            })

        payload.append({
            'id':         each_report.id,
            'uuid':         str(each_report.uuid),
            'date':       each_report.report_date.strftime('%Y-%m-%d'),
            'type':       each_report.report_type,
            'status':     each_report.status.lower(),
            'doctor':     each_report.doctor.profile.user.get_full_name(),
            'parameters': params,
            'notes':      each_report.report_description,
        })

    context = {
            'profile': profile,
            'lab_reports':json.dumps(payload, cls=DjangoJSONEncoder),
            'abnormal_reports': abnormal_reports, 
            'pending_reports': pending_reports,
        }
    
    return render(request, 'pages/patient/lab_report.html', context)

@login_required_with_message(login_url='account:login', message="You need to log in to Download PDF.", only=['patient'])
def lab_report_pdf(request, uuid):
    report = get_object_or_404(LabReport, uuid=uuid)
    html = render_to_string('pages/patient/lab_report_pdf.html', {
        'report': report,
        'parameters': report.parameters.all(),
    })
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="lab_report_{report.uuid}.pdf"'
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error generating PDF', status=500)
    return response


@login_required_with_message(login_url='account:login', message="You need to log in to access your Prescription .", only=['patient'])
def prescriptions(request: HttpRequest):
    if request.method == 'GET':
        """Prescription page view."""
        profile = request.user.profile
        prescriptions = Prescription.objects.select_related('medicine', 'prescribing_doctor').filter(profile=profile)

        active_data = []
        history_data = []
        todayMedications = []

        active_prescriptions = 0
        medicine_to_have = 0

        for pres in prescriptions:
            data = {
                "id": pres.id,
                "name": pres.medicine.name,
                "dosage": pres.update_dosage,
                "frequency": pres.update_frequency,
                "doctor": f"Dr. {pres.prescribing_doctor.profile.user.first_name}",
                "specialty": pres.prescribing_doctor.specialization,
                "startDate": pres.start_date.strftime('%Y-%m-%d'),
                "endDate": pres.end_date.strftime('%Y-%m-%d'),
                "status": pres.status.capitalize(),
            }



            if pres.status == "active":
                data.update({
                    "refillStatus": "Available",  # placeholder: can add logic here
                    "instructions": pres.medicine.instructions,
                    "sideEffects": pres.medicine.side_effects,
                })

                pre_shed: PrescriptionSchedule = PrescriptionSchedule.objects.filter(prescription=pres).all().order_by('time')
                if pre_shed:
                    for each_shed in pre_shed:
                        medicine_to_have+=1
                        todayMedications.append({
                            "name": pres.medicine.name,
                            "time": each_shed.time.strftime('%I:%M %p'),  # 12-hour format with AM/PM
                            "taken": each_shed.had_taken,
                        })

                active_prescriptions+=1
                active_data.append(data)

            else:
                data["reason"] = pres.notes or "Course ended"
                history_data.append(data)

        merged_json_data = {
           'active_data' : json.dumps( active_data, cls=DjangoJSONEncoder),
           'history_data' : json.dumps(history_data, cls=DjangoJSONEncoder),
           'todayMedications' : json.dumps(todayMedications, cls=DjangoJSONEncoder)
        }

        context = {
            'profile': profile,
            'prescriptions': prescriptions,
            'merged_json_data':merged_json_data,
            'active_prescriptions':active_prescriptions,
            'medicine_to_have': medicine_to_have,
            'total_history_data': len(history_data),
        }
        return render(request, 'pages/patient/prescriptions.html', context)

@login_required_with_message(login_url='account:login', message="You need to log in to access Profile page.", only=['patient'])
def p_profile(request: HttpRequest):
    """Patient profile page view."""

    if request.method == 'GET':
        # Fetch the user's profile information
        profile: Profile = request.user.profile

        # Pass the profile information to the template
        context = {
            'profile': profile,
        }

        return render(request, 'pages/patient/profile.html', context)
    
    else:
        try:
            profile: Profile = request.user.profile
            user: User = request.user

            # Handle profile picture
            profile_pic = request.FILES.get('profileImage')  
            if profile_pic:
                profile.profile_pic = profile_pic

            # Personal Info
            user.first_name = request.POST.get('full_name', '').strip()
            # user.email = request.POST.get('email', '').strip()
            profile.ph_number = request.POST.get('phone_number', '').strip()
            profile.address = request.POST.get('address', '').strip()

            dob_str = request.POST.get('date_of_birth', '').strip()
            if dob_str:
                try:
                    profile.date_of_birth = datetime.strptime(dob_str, '%Y-%m-%d').date()
                except ValueError:
                    messages.error(request, "Invalid date format for Date of Birth.")
                    return redirect('patient:profile')

            # Medical Info
            medical_info, created = MedicalInfo.objects.get_or_create(profile=profile)
            medical_info.blood_group = request.POST.get('blood_group', '').strip()
            medical_info.medical_conditions = request.POST.get('medicalConditions', '').strip()  # fixed name
            medical_info.allergies = request.POST.get('allergies', '').strip()  # make sure this field exists
            medical_info.on_going_medications = request.POST.get('medicines_on', '').strip()  # make sure this field exists

            # Emergency Contact
            medical_info.emg_contact_name = request.POST.get('emergency_contact_name', '').strip()
            medical_info.emg_contact_number = request.POST.get('emergency_contact_number', '').strip()
            medical_info.emg_contact_relation = request.POST.get('emergency_contact_relationship', '').strip()
            medical_info.emg_contact_address = request.POST.get('emergency_contact_address', '').strip()

            # Notification Settings
            profile.email_notification = 'emailNotifications' in request.POST
            profile.reminders = 'appointmentReminders' in request.POST


            medical_info.save()
            request.user.save()
            profile.save()

            messages.success(request, "Profile updated successfully.")
        except Exception as e:
            print(f"Error updating profile: {e}")
            messages.error(request, "An error occurred while updating your profile.")

        return redirect('patient:profile')


@login_required_with_message(login_url='account:login', message="You need to log in to View your Activities.", only=['patient'])
def p_activities(request: HttpRequest):
    """Patient activities page view."""
    profile: Profile = request.user.profile
    activities = ActivityLog.objects.filter(profile=profile).order_by('-timestamp')

    context = {
        'profile': profile,
        'activities': activities,
    }
    return render(request, 'pages/patient/activities.html', context)

# --------------------------------------- Adding Logic on each pages ------------------------------------------------------------
